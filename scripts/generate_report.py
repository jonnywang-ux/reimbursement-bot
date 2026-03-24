#!/usr/bin/env python3
"""
Generate a Budget Realization report (CSV + Excel) from extracted expense data.

Usage:
    python generate_report.py <expenses_json> <output_dir> [claimant_name]
"""

import json
import os
import sys
from datetime import date

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def generate_report(expenses_json_path, output_dir, claimant_name="Daniel", reimbursement_purpose=""):
    with open(expenses_json_path, "r") as f:
        expenses = json.load(f)

    os.makedirs(output_dir, exist_ok=True)

    columns = [
        "No.",
        "Date",
        "User",
        "Expense Type",
        "Vendor Name",
        "Description",
        "Currency",
        "Price/Unit",
        "Total Nights",
        "Total Units",
        "Total Amount (Original)",
        "MAS Rate Date",
        "FX Rate (SGD per unit)",
        "Total Amount (SGD)",
        "FX Rate (USD per unit)",
        "Total Amount (USD)",
        "Paid By",
    ]

    rows = []
    total_sgd = 0.0
    total_usd = 0.0
    has_fallback = False

    for i, exp in enumerate(expenses, start=1):
        sgd = float(exp.get("total_amount_sgd", 0))
        usd = float(exp.get("total_amount_usd", 0))
        total_sgd += sgd
        total_usd += usd

        if exp.get("fx_source") == "fallback":
            has_fallback = True

        rows.append([
            i,
            exp.get("date", ""),
            exp.get("user", claimant_name),
            exp.get("expense_type", exp.get("category", "")),
            exp.get("vendor_name", ""),
            exp.get("description", ""),
            exp.get("currency", ""),
            f'{float(exp.get("price_per_unit", exp.get("amount", 0))):.2f}',
            exp.get("total_nights", "") if exp.get("total_nights") is not None else "",
            exp.get("total_units", 1) if exp.get("total_units") is not None else 1,
            f'{float(exp.get("total_amount_original", exp.get("amount", 0))):.2f}',
            exp.get("mas_rate_date", exp.get("rate_date", "")),
            f'{float(exp.get("fx_rate_sgd", 0)):.6f}',
            f'{sgd:.2f}',
            f'{float(exp.get("fx_rate_usd", 0)):.6f}',
            f'{usd:.2f}',
            exp.get("paid_by", claimant_name),
        ])

    # Build month/year from first expense date for filename
    first_date = expenses[0].get("date", "") if expenses else ""
    try:
        from datetime import datetime
        dt = datetime.strptime(first_date, "%Y-%m-%d")
        month_year = dt.strftime("%B_%Y")
    except (ValueError, TypeError):
        month_year = date.today().strftime("%B_%Y")

    # Blank row then TOTAL row
    blank = [""] * len(columns)
    total_row = [""] * len(columns)
    total_row[0] = ""
    total_row[12] = "TOTAL"   # FX Rate (SGD) column header repurposed as label
    total_row[13] = f"{total_sgd:.2f}"
    total_row[15] = f"{total_usd:.2f}"

    rows.append(blank)
    rows.append(total_row)

    # Fallback footnote if needed
    if has_fallback:
        footnote = [""] * len(columns)
        footnote[0] = "⚠️ Some rates sourced from open.er-api.com fallback, not official MAS rates."
        rows.append(footnote)

    df = pd.DataFrame(rows, columns=columns)

    safe_name = claimant_name.replace(" ", "_")
    base_name = f"{safe_name}_{month_year}_Reimbursement"

    csv_path = os.path.join(output_dir, f"{base_name}.csv")
    df.to_csv(csv_path, index=False)

    xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        # ── Page 1: Reimbursement Purpose ────────────────────────────────────
        # Determine date range from expenses
        all_dates = [exp.get("date", "") for exp in expenses if exp.get("date")]
        date_range = ""
        if all_dates:
            sorted_dates = sorted(all_dates)
            if sorted_dates[0] == sorted_dates[-1]:
                date_range = sorted_dates[0]
            else:
                date_range = f"{sorted_dates[0]} to {sorted_dates[-1]}"

        purpose_rows = [
            ["Gunung Capital — Reimbursement Request"],
            [],
            ["Claimant", claimant_name],
            ["Date Range", date_range],
            ["Total (SGD)", f"{total_sgd:.2f}"],
            ["Total (USD)", f"{total_usd:.2f}"],
            [],
            ["Purpose / Description"],
            [reimbursement_purpose or "(Not provided)"],
        ]
        purpose_df = pd.DataFrame(purpose_rows)
        purpose_df.to_excel(writer, index=False, header=False, sheet_name="Cover")
        ws_cover = writer.sheets["Cover"]

        # Style the cover sheet
        title_font = Font(size=16, bold=True, color="1F4E79")
        ws_cover["A1"].font = title_font
        ws_cover["A1"].alignment = Alignment(horizontal="left")
        label_font = Font(bold=True)
        for row_num in [3, 4, 5, 6, 9]:
            ws_cover.cell(row=row_num, column=1).font = label_font
        ws_cover.column_dimensions["A"].width = 22
        ws_cover.column_dimensions["B"].width = 50

        # ── Page 2: Budget Realization (existing layout) ─────────────────────
        df.to_excel(writer, index=False, sheet_name="Budget Realization")
        ws = writer.sheets["Budget Realization"]

        # Style header row
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Style TOTAL row (second to last data row — blank row before it)
        total_row_idx = len(rows)  # 1-indexed + 1 for header = len(rows)+1, but blank added too
        # Find TOTAL row: it's at position len(expenses)+3 (1-indexed: header + expenses + blank + total)
        total_excel_row = len(expenses) + 3
        for cell in ws[total_excel_row]:
            cell.font = Font(bold=True)

        # Auto-width columns
        for idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(idx)
            col_values = [str(col_name)] + [str(r[idx - 1]) for r in rows]
            max_len = max(len(v) for v in col_values) + 2
            ws.column_dimensions[col_letter].width = min(max_len, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

    return csv_path, xlsx_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    name = sys.argv[3] if len(sys.argv) > 3 else "Daniel"
    purpose = sys.argv[4] if len(sys.argv) > 4 else ""
    csv_out, xlsx_out = generate_report(sys.argv[1], sys.argv[2], name, purpose)
    print(f"CSV:  {csv_out}")
    print(f"XLSX: {xlsx_out}")
